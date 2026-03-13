from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = "/Users/jackvoller/desk_booking/CT600_v1_993_to_v1_994_differences_with_boxes.xlsx"

rows = [
    ("CT.xsd", "Schema metadata", "xsd:schema@version", "N/A", "Updated", "Schema version changed 1.993 -> 1.994", "Update metadata/version handling"),
    ("CT.xsd", "Schema metadata", "gms:Modified", "N/A", "Updated", "Modified date changed 2024-11-15 -> 2025-10-01", "No calculation impact"),
    ("CT.xsd", "Schema metadata", "gms:Copyright", "N/A", "Updated", "Copyright changed to 2001-2025", "No calculation impact"),

    ("CT.xsd + CT.sch", "Charity", "Charity/InformationRequired/Income/LegacyIncome", "E88", "Added", "New legacy income field", "Add mapping, input and validation"),
    ("CT.xsd + CT.sch", "Charity", "Charity/InformationRequired/Assets/LegacyPayments", "E195A-E200", "Added", "New repeating legacy payments block", "Add repeating serialization and totals"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Payment/Forename", "E195A", "Added", "New payment forename", "Map field"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Payment/Surname", "E195B", "Added", "New payment surname", "Map field"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Payment/Address", "E195C", "Added", "New payment address", "Map field"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Payment/Postcode", "E195D", "Added", "Optional postcode", "Add postcode/overseas rule"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Payment/Overseas", "E195E", "Added", "Optional overseas indicator", "Add postcode/overseas rule"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Payment/Date", "E195F", "Added", "Payment date <= period end [35]", "Add date bound check"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Payment/Amount", "E195G", "Added", "Payment amount", "Map field"),
    ("CT.xsd + CT.sch", "Charity", ".../LegacyPayments/Total", "E200", "Added", "Total equals sum of E195G", "Add cross-field sum check"),
    ("CT.sch", "Charity", "Charity/InformationRequired/Income/Total", "E90", "Updated validation", "E90 now includes E88", "Update income total formula"),

    ("CT.xsd + CT.sch", "R&D Step3", "ResearchAndDevelopment/Step3/ExceptionS1112E", "L71", "Added", "New exception field", "Add mapping and dependencies"),
    ("CT.xsd + CT.sch", "R&D Step3", ".../TotalExpenditureOnEPWsFromAndSubcontractingToConnectedPersons", "L71A", "Added", "New connected-person EPW/subcontracting field", "Add mapping and cap logic"),
    ("CT.xsd + CT.sch", "R&D Step3", ".../PAYENICsForWhichTheCompanyIsLiableInThisAP", "L72", "Added", "New PAYE/NIC amount", "Add mapping and formula integration"),
    ("CT.xsd + CT.sch", "R&D Step3", ".../EmployerPAYEreference", "L72A", "Added", "Company PAYE references", "Add structure and format checks"),
    ("CT.xsd + CT.sch", "R&D Step3", ".../RelevantPAYENICsLiabilityOfConnectedCompanies", "L73", "Added", "Connected companies PAYE/NIC", "Add mapping and formula integration"),
    ("CT.xsd + CT.sch", "R&D Step3", ".../ConnectedCompaniesEmployerPAYEreference", "L73A", "Added", "Connected company PAYE references", "Add structure and format checks"),
    ("CT.sch", "R&D Step3", ".../TotalRelevantExpenditureOnRandDWorkersPAYEandNationalInsuranceContributions", "L75", "Updated validation", "Post-2024 cap logic changed", "Update cap formula"),
    ("CT.sch", "R&D Step3", ".../Step3RestrictionCarriedForwardToNextAP", "L80", "Updated validation", "Carry-forward rule updated for exception paths", "Update L80 logic"),

    ("CT.sch", "Tax reconciliation", "TaxReconciliation/CreativeCredit", "540, 96", "Added validation", "540 allowed only if 96=yes", "Gate by CT600P flag"),
    ("CT.sch", "Tax reconciliation", "TaxReconciliation/AVECandVGEC", "541, 96", "Added validation", "541 allowed only if 96=yes", "Gate by CT600P flag"),
    ("CT.sch", "Enhanced expenditure", "EnhancedExpenditure/CreativesCoreExpenditure", "663, 96", "Added validation", "663 allowed only if 96=yes", "Gate by CT600P flag"),
    ("CT.sch", "Enhanced expenditure", "EnhancedExpenditure/CreativeEnhancedExpenditure", "665, 96", "Added validation", "665 allowed only if 96=yes", "Gate by CT600P flag"),

    ("CT.xsd + CT.sch", "AVEC", ".../Film/UKrelevantGlobalExpenditure", "P5B", "Added", "New UK relevant global expenditure", "Add mapping and <=P5A rule"),
    ("CT.xsd + CT.sch", "AVEC", ".../HighEndTV/UKrelevantGlobalExpenditure", "P10B", "Added", "New UK relevant global expenditure", "Add mapping and <=P10A rule"),
    ("CT.xsd + CT.sch", "AVEC", ".../ChildrensTV/UKrelevantGlobalExpenditure", "P15B", "Added", "New UK relevant global expenditure", "Add mapping and <=P15A rule"),
    ("CT.xsd + CT.sch", "AVEC", ".../Animation/UKrelevantGlobalExpenditure", "P20B", "Added", "New UK relevant global expenditure", "Add mapping and <=P20A rule"),
    ("CT.xsd + CT.sch", "AVEC", ".../IndependentFilm/UKrelevantGlobalExpenditure", "P25B", "Added", "New UK relevant global expenditure", "Add mapping and <=P25A rule"),
    ("CT.xsd + CT.sch", "AVEC totals", ".../TotalAudioVisual/TotalUKrelevantGlobalExpenditure", "P30B", "Added", "New total UK relevant global expenditure", "Add total from P5B/P10B/P15B/P20B/P25B"),
    ("CT.xsd + CT.sch", "AVEC totals", ".../TotalAudioVisual/TotalQualifyingExpenditure", "P30C", "Updated mapping text", "Box reference shifted to P30C", "Update mapping references"),
    ("CT.xsd + CT.sch", "AVEC totals", ".../TotalAudioVisual/TotalExpenditureCreditClaimed", "P30D", "Updated mapping text", "Box reference shifted to P30D", "Update mapping references"),
    ("CT.xsd + CT.sch", "AVEC totals", ".../TotalAudioVisual/TotalAdditionalCreditForVisualEffects", "P30E", "Added", "New total additional VFX credit", "Add total and linkage to P81"),

    ("CT.xsd + CT.sch", "VGEC", ".../VideoGames/UKrelevantGlobalExpenditure", "P35B", "Added", "New UK relevant global expenditure", "Add mapping and <=P35A rule"),
    ("CT.xsd + CT.sch", "VGEC totals", ".../TotalVideoGames/TotalUKrelevantGlobalExpenditure", "P45B", "Added", "New total UK relevant global expenditure", "Add equality to P35B"),
    ("CT.xsd + CT.sch", "VGEC totals", ".../TotalVideoGames/TotalQualifyingExpenditure", "P45C", "Updated mapping text", "Box reference shifted to P45C", "Update mapping references"),
    ("CT.xsd + CT.sch", "VGEC totals", ".../TotalVideoGames/TotalExpenditureCreditClaimed", "P45D", "Updated mapping text", "Box reference shifted to P45D", "Update mapping references"),

    ("CT.xsd + CT.sch", "Creative Step1", "CreativeIndustries/Step1/AdditionalAVECcredit", "P81", "Added", "New additional AVEC credit field", "Add mapping and equality to P30E"),
    ("CT.sch", "Creative Step1", "CreativeIndustries/Step1/TotalAVECandVGECforTheAccountingPeriod", "P95", "Updated validation", "P95 now sums P80+P81+P90", "Update total formula"),

    ("CT.xsd + CT.sch", "Film/TV/Animation/VG tax relief", ".../Film/UKcoreExpenditure", "P260B", "Added", "New UK core expenditure", "Add mapping and <=P260A rule"),
    ("CT.xsd + CT.sch", "Film/TV/Animation/VG tax relief", ".../HighEndTV/UKcoreExpenditure", "P265B", "Added", "New UK core expenditure", "Add mapping and <=P265A rule"),
    ("CT.xsd + CT.sch", "Film/TV/Animation/VG tax relief", ".../ChildrensTV/UKcoreExpenditure", "P270B", "Added", "New UK core expenditure", "Add mapping and <=P270A rule"),
    ("CT.xsd + CT.sch", "Film/TV/Animation/VG tax relief", ".../Animation/UKcoreExpenditure", "P275B", "Added", "New UK core expenditure", "Add mapping and <=P275A rule"),
    ("CT.xsd + CT.sch", "Film/TV/Animation/VG tax relief", ".../VideoGames/UKcoreExpenditure", "P280B", "Added", "New UK core expenditure", "Add mapping and <=P280A rule"),
    ("CT.xsd + CT.sch", "Film/TV/Animation/VG totals", ".../TotalAudioVisualAndVideoGames/TotalUKcoreExpenditure", "P285B", "Added", "New total UK core expenditure", "Add total from P260B/P265B/P270B/P275B/P280B"),
    ("CT.sch", "Film/TV/Animation/VG totals", ".../TotalAudioVisualAndVideoGames/TotalAdditionalDeduction", "P285C", "Updated mapping text", "Box reference shifted to P285C", "Update mapping references"),
    ("CT.sch", "Film/TV/Animation/VG totals", ".../TotalAudioVisualAndVideoGames/TotalLossesSurrenderedForTaxCredit", "P285D", "Updated mapping text", "Box reference shifted to P285D", "Update mapping references"),
    ("CT.sch", "Film/TV/Animation/VG totals", ".../TotalAudioVisualAndVideoGames/TotalTaxCreditClaimed", "P285E", "Updated mapping text", "Box reference shifted to P285E", "Update mapping references"),

    ("CT.xsd + CT.sch", "Cultural reliefs", ".../Theatre/UKcoreExpenditure", "P290B", "Added", "New UK core expenditure", "Add mapping and <=P290A rule"),
    ("CT.xsd + CT.sch", "Cultural reliefs", ".../Orchestra/UKcoreExpenditure", "P295B", "Added", "New UK core expenditure", "Add mapping and <=P295A rule"),
    ("CT.xsd + CT.sch", "Cultural reliefs", ".../MuseumGalleryExhibition/UKcoreExpenditure", "P300B", "Added", "New UK core expenditure", "Add mapping and <=P300A rule"),
    ("CT.xsd + CT.sch", "Cultural totals", ".../TotalCulturalReliefs/TotalUKcoreExpenditure", "P305B", "Added", "New total UK core expenditure", "Add total from P290B/P295B/P300B"),
    ("CT.sch", "Cultural totals", ".../TotalCulturalReliefs/TotalAdditionalDeduction", "P305C", "Updated mapping text", "Box reference shifted to P305C", "Update mapping references"),
    ("CT.sch", "Cultural totals", ".../TotalCulturalReliefs/TotalLossesSurrenderedForTaxCredit", "P305D", "Updated mapping text", "Box reference shifted to P305D", "Update mapping references"),
    ("CT.sch", "Cultural totals", ".../TotalCulturalReliefs/TotalTaxCreditClaimed", "P305E", "Updated mapping text", "Box reference shifted to P305E", "Update mapping references"),

    ("CT.sch", "Company type/rate", "assert a_F041.6", "4, 390", "Updated validation text", "Error text now references [390] (was [340])", "Update help/error text mapping"),
    ("CT.sch", "Allowances and charges", "AllowancesAndCharges/EnterpriseZones", "721/722, 30", "Added validation", "Component not allowed if [30] on/after 2024-04-01", "Add date gate"),
    ("CT.sch", "Not Included", "NotIncluded/EnterpriseZones", "746/747, 30", "Added validation", "Component not allowed if [30] on/after 2024-04-01", "Add date gate"),
    ("CT.sch", "Allowances and charges", "AllowancesAndCharges/ZeroEmissionsGoodsVehicles", "723/724, 30", "Updated validation", "Whole component disallowed from 2025-04-01", "Update date gate"),
    ("CT.sch", "Not Included", "NotIncluded/ZeroEmissionsGoodsVehicles", "748/749, 30", "Updated validation", "Whole component disallowed from 2025-04-01", "Update date gate"),
]

wb = Workbook()
ws = wb.active
ws.title = "Differences With Boxes"
headers = [
    "Source",
    "Section",
    "Element/Tag",
    "Box Number(s)",
    "Change Type",
    "Description of Change",
    "What You Need To Update",
]
ws.append(headers)
for row in rows:
    ws.append(row)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx in range(1, len(headers) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        value = row[0].value
        length = len(str(value)) if value is not None else 0
        if length > max_len:
            max_len = length
    ws.column_dimensions[col_letter].width = min(max_len + 2, 68)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

ws.freeze_panes = "A2"

summary = wb.create_sheet("Summary")
summary["A1"] = "CT600 v1.993 -> v1.994 Differences (XSD + Schematron)"
summary["A1"].font = Font(bold=True, size=14)
summary["A3"] = "Generated on"
summary["B3"] = str(date.today())
summary["A4"] = "Total tracked rows"
summary["B4"] = len(rows)
summary["A6"] = "Note"
summary["B6"] = "Workbook combines structural changes (XSD) with validation/box references (Schematron). Box references are taken from assertion texts in the provided diff."
summary.column_dimensions["A"].width = 26
summary.column_dimensions["B"].width = 130
summary["B6"].alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print(OUT)
